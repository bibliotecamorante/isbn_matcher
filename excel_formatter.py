# -*- coding: utf-8 -*-
"""
excel_formatter.py - Formattazione file Excel output
Applica lo stile dell'app Sebina Plus: colori, larghezze, altezze
Aggiornato per supportare localizzazione
"""
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter
from typing import Callable, Optional
from config import AppConfig
from localization import Translations


def formatta_excel_isbn(
    filepath: Path,
    config: AppConfig,
    log_callback: Callable[[str, str], None],
    progress_callback: Optional[Callable[[int, int], None]] = None,
    t: Optional[Translations] = None
) -> None:
    """
    Formatta il file Excel di output con lo stile Sebina Plus.
    
    Stile applicato:
    - Header: sfondo azzurro (#8DBEE3), testo bold, wrap text, abbreviazioni
    - Righe: altezza 30, wrap text, allineamento top-left
    - Colonne: larghezza specifica per colonna o default 15 caratteri
    - Colonna ISBN: larghezza 18 caratteri
    - Freeze panes: prima riga (header)
    - Zoom: 110%
    
    Args:
        filepath: File Excel da formattare
        config: Configurazione applicazione
        log_callback: Funzione per logging (message, level)
        progress_callback: Funzione per progress bar (current, total)
        t: Traduzioni (opzionale)
    """
    start_msg = t.proc_applying_format if t else "Avvio formattazione Excel..."
    log_callback(start_msg, "INFO")
    
    try:
        wb = load_workbook(str(filepath))
        
        # Registra stili globali
        _registra_stili_globali(wb, config)
        
        for ws in wb.worksheets:
            if ws.title.lower() == config.SHEET_PARAMETRI:
                continue
            
            sheet_msg = f"  {t.proc_formatting_sheet if t else 'Formattazione foglio'}: {ws.title}"
            log_callback(sheet_msg, "INFO")
            
            # Setup pagina
            _setup_pagina(ws, config)
            
            max_row = ws.max_row
            max_col = ws.max_column
            
            # Applica abbreviazioni agli header
            _applica_abbreviazioni(ws, max_col, config)
            
            # Formatta header (riga 1)
            _formatta_header(ws, max_col, config)
            
            # Formatta colonne (larghezze)
            _formatta_larghezze_colonne(ws, max_col, config)
            
            # Formatta righe dati
            _formatta_righe_dati(ws, max_row, max_col, config)
            
            if progress_callback:
                progress_callback(80, 100)
        
        if progress_callback:
            progress_callback(100, 100)
        
        save_msg = t.proc_saving_format if t else "Salvataggio formattazione..."
        log_callback(save_msg, "INFO")
        wb.save(str(filepath))
        
        complete_msg = t.proc_format_complete if t else "✅ Formattazione completata"
        log_callback(complete_msg, "SUCCESS")
        
    except PermissionError:
        if t:
            error_msg = t.error_file_open_excel.format(filename=filepath.name)
        else:
            error_msg = f"Il file '{filepath.name}' è aperto in Excel.\nChiudilo e riprova l'operazione."
        raise PermissionError(error_msg)
    except Exception as e:
        error_msg = f"{t.error_formatting if t else '❌ Errore formattazione'}: {str(e)}"
        log_callback(error_msg, "ERROR")
        raise


def _registra_stili_globali(wb, config: AppConfig) -> None:
    """Registra stili riutilizzabili (come Sebina Plus)"""
    stili = {
        "header_style": NamedStyle(
            name="header_style",
            fill=PatternFill(
                start_color=config.EXCEL_COLOR_HEADER, 
                end_color=config.EXCEL_COLOR_HEADER, 
                fill_type='solid'
            ),
            alignment=Alignment(wrap_text=True, vertical='top', horizontal='left'),
            font=Font(bold=True, color=config.EXCEL_COLOR_HEADER_FONT)
        ),
        "normal_style": NamedStyle(
            name="normal_style",
            alignment=Alignment(wrap_text=True, vertical='top', horizontal='left')
        )
    }
    
    for style in stili.values():
        if style.name not in wb.named_styles:
            wb.add_named_style(style)


def _setup_pagina(ws, config: AppConfig) -> None:
    """Setup configurazione pagina (come Sebina Plus)"""
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.freeze_panes = 'A2'  # Blocca header
    ws.sheet_view.zoomScale = config.EXCEL_ZOOM


def _applica_abbreviazioni(ws, max_col: int, config: AppConfig) -> None:
    """
    Applica abbreviazioni agli header delle colonne.
    """
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        if cell.value:
            header_lower = str(cell.value).lower().strip()
            
            # Cerca se c'è un'abbreviazione configurata
            if header_lower in config.ABBREV:
                cell.value = config.ABBREV[header_lower]


def _formatta_header(ws, max_col: int, config: AppConfig) -> None:
    """
    Formatta la riga di intestazione (riga 1).
    Stile: sfondo azzurro, testo bold, altezza configurabile.
    """
    ws.row_dimensions[1].height = config.EXCEL_ROW_HEIGHT_HEADER
    
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.style = "header_style"


def _formatta_larghezze_colonne(ws, max_col: int, config: AppConfig) -> None:
    """
    Imposta larghezza colonne in base alla configurazione.
    - Colonne con larghezze specifiche: usa config.LARGHEZZE
    - Colonne ISBN: 18 caratteri
    - Altre colonne: 15 caratteri (default)
    """
    from utils import is_isbn_column_name  # Import locale per evitare dipendenze circolari
    
    # Applica larghezze
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        header_val = ws.cell(row=1, column=col_idx).value
        
        if header_val and is_isbn_column_name(str(header_val), config):
            # Colonne ISBN
            ws.column_dimensions[col_letter].width = config.EXCEL_COLUMN_WIDTH_ISBN
        elif header_val:
            # Cerca larghezza specifica
            header_lower = str(header_val).lower().strip()
            
            if header_lower in config.LARGHEZZE:
                ws.column_dimensions[col_letter].width = config.LARGHEZZE[header_lower]
            else:
                ws.column_dimensions[col_letter].width = config.LARGHEZZE_DEFAULT
        else:
            ws.column_dimensions[col_letter].width = config.LARGHEZZE_DEFAULT


def _formatta_righe_dati(ws, max_row: int, max_col: int, config: AppConfig) -> None:
    """
    Formatta righe dati (dalla riga 2 in poi).
    - Altezza: configurabile
    - Stile: wrap text, allineamento top-left
    
    OTTIMIZZATO: Usa iter_rows per performance migliori
    """
    if max_row < 2:  # Nessuna riga dati
        return
    
    # Applica stile e altezza in un solo loop ottimizzato
    for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
        # Imposta altezza della riga
        ws.row_dimensions[row[0].row].height = config.EXCEL_ROW_HEIGHT_DATA
        
        # Applica stile a tutte le celle della riga
        for cell in row:
            cell.style = "normal_style"
